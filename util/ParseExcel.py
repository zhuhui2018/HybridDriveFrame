# -*-coding:utf-8 -*-
# @Author : Zhigang

import openpyxl
from openpyxl.styles import Border,Side,Font
from datetime import datetime

class ParseExcel(object):

    def __init__(self):
        self.workbook=None
        self.excelFile=None
        self.font=Font(color=None)   # 设置字体的颜色
        self.RGBDict = {"red": "00FF0000", "green": "0000FF00"}

    def loadWorkBook(self,excelPathAndName):
        "将excel文件加载到内存，并获取其workbook对象"
        try:
            self.workbook=openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        else:
            self.excelFile=excelPathAndName
            return self.workbook

    def getSheetByName(self,sheetName):
        "通过sheet名获取sheet对象"
        try:
            sheet=self.workbook[sheetName]
        except Exception as e:
            raise e
        else:
            return sheet

    def getSheetByIndex(self,sheetIndex):
        "通过sheet的索引号来获取sheet对象"
        sheetNameList=self.workbook.sheetnames
        try:
            sheet=self.workbook[sheetNameList[sheetIndex]]
        except Exception as e:
            raise e
        else:
            return sheet

    def getRowsNumber(self,sheet):
        "获取sheet中有数据区域的结束行号"
        return sheet.max_row

    def getColsNumber(self,sheet):
        "获取sheet中有数据区域的结束列号"
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        "获取sheet中有数据区域的开始行号"
        return sheet.min_row

    def getStartColNumber(self,sheet):
        "获取sheet中有数据区域的开始列号"
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        "获取sheet中的某一行，返回的是这一行所有的数据内容组成的tuple，"
        # 下标从1开始，rowNo=1表示读取第一行
        try:
            rows=[]
            for row in sheet.iter_rows():
                rows.append(row)
            # for cell in rows[rowNo-1]:
            #     print (cell.value)
            return rows[rowNo-1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):
        "获取sheet中某一列，返回的是这一列所有的数据内容组成tuple."
        # 下标从1开始，sheet.columns[1]表示第一列
        try:
            cols=[]
            for col in sheet.iter_cols():
                cols.append(col)
            # for cell in cols[colNo-1]:
            #     print (cell.value)
            return cols[colNo-1]
        except Exception as e:
            raise e

    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        "根据单元格所在的位置索引获取该单元格中的值，下标从1开始，"
        # sheet.cell(row=1,column=1).value 表示excel中第一行第一列的值
        if coordinate!=None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        "获取单元格的对象，可通过单元格所在位置的数字索引，也可通过excel单元格中的编码及坐标"
        # getCellOfObject(sheet,coordinate="B2") or getCellOfObject(sheet,rowNo=2,colsNo=3)
        if coordinate!=None:
            try:
                return sheet[coordinate]
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row=rowNo,column=colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCell(self,sheet,content,coordinate=None,rowNo=None,colsNo=None,style=None):
        "通过单元格在excel中的编码坐标或者数字索引坐标向单元格中写入数据"
        # 下标从1开始，参数style表示字体的颜色的名字，比如red,green
        if coordinate is not None:
            try:
                cell=self.getCellOfObject(sheet,coordinate)
                cell.value=content
                if style is not None:
                    cell.font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                cell=self.getCellOfObject(sheet,rowNo=rowNo,colsNo=colsNo)
                cell.value=content
                if style is not None:
                    # print (style)
                    # print (self.RGBDict[style])
                    cell.font=Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCellCurrentTime(self,sheet,coordinate=None,rowNo=None,colsNo=None):
        "写入当前的时间，下标从1开始"
        # currentTime=time.strftime("%Y-%m-%d %H:%M:%S")
        currentTime=datetime.now()
        if coordinate is not None:
            try:
                cell=self.getCellOfObject(sheet,coordinate)
                cell.value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colsNo is not None:
            try:
                cell=self.getCellOfObject(sheet,rowNo=rowNo,colsNo=colsNo)
                cell.value=currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

if __name__=="__main__":
    p=ParseExcel()
    p.loadWorkBook("C:\\Users\\zhigang\\Desktop\\126邮箱创建联系人并发邮件.xlsx")
    # print (p.getSheetByName("登录"))
    # print (p.getRowsNumber(p.getSheetByIndex(4)))
    # print (p.getColsNumber(p.getSheetByIndex(4)))
    # print (p.getStartRowNumber(p.getSheetByName("登录")))
    # print (p.getStartColNumber(p.getSheetByName("发邮件")))
    # print (p.getRow(p.getSheetByIndex(4),1))
    # print (p.getColumn(p.getSheetByName("登录"),2))
    # print (p.getCellOfValue(p.getSheetByName("登录"),coordinate="B2"))
    # print (p.getCellOfValue(p.getSheetByName("登录"),rowNo=2,colsNo=5))
    # print (p.getCellOfObject(p.getSheetByName("登录"),coordinate="B2"))
    # print (p.getCellOfObject(p.getSheetByName("登录"),rowNo=2,colsNo=5))
    sheet=p.getSheetByIndex(4)
    print (sheet.title)
    p.writeCell(sheet,coordinate="H6",content="您好")
    p.writeCellCurrentTime(sheet,coordinate="H3")
    p.writeCell(sheet,rowNo=4,colsNo=8,content="新年好")
    p.writeCellCurrentTime(sheet,rowNo=4,colsNo=9)